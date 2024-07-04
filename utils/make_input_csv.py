import os
from tqdm import tqdm
import openpyxl
import pandas as pd

class MAKE_CLASSIFICATION_DATASET:
    def __init__(self, dataset_path='data', project_name='project'):
        self.dataset_path = dataset_path
        self.project_name = project_name
        self.dataset = dataset_path.split('/')[-1]
        self.project_path = os.path.join(project_name, self.dataset + '_dataset.xlsx')
        if not os.path.exists(project_name+'/temp'):os.makedirs(project_name+'/temp')
    
    def make_input_excel(self):
        # 創建一個新的 Excel 工作簿
        if not os.path.exists(self.project_path):
            openpyxl.Workbook().save(self.project_path)
        workbook = openpyxl.load_workbook(self.project_path, data_only=True)
        data_sheet = workbook.active
        data_sheet.title = "classification_dataset"
        
        img_count,num_class = 0,0
        
        # 計算總圖片數
        for category in os.listdir(self.dataset_path):
            num_class += 1
            img_count += len(os.listdir(os.path.join(self.dataset_path, category)))
        
        mission = tqdm(total=img_count)
        
        # 寫入標題行
        data_sheet.append(["category", "set", "img_path"])
        
        # 創建一個新的工作表寫入總結信息
        summary_sheet = workbook.create_sheet(title="classification_dataset_Summary")
        summary_sheet.append(["num_class",num_class])
        summary_sheet.append(['Category', 'Set', 'Count'])
        
        for category_id, category in enumerate(os.listdir(self.dataset_path)):
            print(category_id, category)
            
            check_calculation_map = {}
            for count, img in enumerate(os.listdir(os.path.join(self.dataset_path, category))):
                img_path = os.path.join(self.dataset_path, category, img)
                data_sheet.append([category_id, count % 5, img_path])
                
                check_calculation_map.setdefault(count % 5, 0)
                check_calculation_map[count % 5] += 1
                mission.update()
            
            # 將每個類別的總結信息寫入到總結工作表中
            for set_num, count in check_calculation_map.items():
                summary_sheet.append([category, set_num, count])
            summary_sheet.append([])
        
        # 保存 Excel 文件
        workbook.save(self.project_path)
        
        print("make_input_excel end")
        return 'make_input_excel end'

    def save_sheet_as_csv(self, sheet_name, csv_file_path):
        # 讀取 Excel 文件中的特定工作表
        df = pd.read_excel(self.project_path, sheet_name=sheet_name)

        # 將 DataFrame 存成 CSV 文件
        df.to_csv(csv_file_path, index=False)
        
        print(f'Saved {sheet_name} as {csv_file_path}')

# # 使用範例
# handler = MAKE_CLASSIFICATION_DATASET('cats_and_dogs', 'CAT_DOG')
# handler.make_input_excel()
# handler.save_sheet_as_csv('classification_dataset', 'CAT_DOG/temp/classification_dataset.csv')
# handler.save_sheet_as_csv('classification_dataset_Summary', 'CAT_DOG/temp/classification_dataset_detail.csv')


class MAKE_INFERENCE_DATASET:
    def __init__(self, dataset_path=str, project_name=str):
        self.dataset_path = dataset_path
        self.project_name = project_name
        self.dataset = dataset_path.split('/')[-1]
        self.project_path = os.path.join(project_name, self.dataset + '_dataset.xlsx')
        if not os.path.exists(project_name+'/temp'):os.makedirs(project_name+'/temp')
    
    def make_input_excel(self):
        # 創建一個新的 Excel 工作簿
        #workbook = openpyxl.Workbook()

        if not os.path.exists(self.project_path):
            openpyxl.Workbook().save(self.project_path)
        workbook = openpyxl.load_workbook(self.project_path, data_only=True)
        # data_sheet = workbook.active
        # data_sheet.title = 
        workbook.create_sheet('inference_dataset')
        
        img_count,num_class = 0,0
        
        # 計算總圖片數
        for category in os.listdir(self.dataset_path):
            num_class += 1
            img_count += len(os.listdir(os.path.join(self.dataset_path, category)))
        
        mission = tqdm(total=img_count)
        
        # 寫入標題行
        workbook['inference_dataset'].append(["category", "img_path"])
        
        # 創建一個新的工作表寫入總結信息
        summary_sheet = openpyxl.load_workbook(self.project_path, data_only=True).create_sheet(title="inference_dataset_Summary")
        summary_sheet.append(["num_class",num_class])
        summary_sheet.append([])
        summary_sheet.append(['Category', 'Count'])
        for category_id, category in enumerate(os.listdir(self.dataset_path)):
            print(category_id, category)
            
            check_calculation_map = {}
            for count, img in enumerate(os.listdir(os.path.join(self.dataset_path, category))):
                img_path = os.path.join(self.dataset_path, category, img)
                workbook['inference_dataset'].append([category_id, img_path])
                
                check_calculation_map.setdefault(count % 1, 0)
                check_calculation_map[count % 1] += 1
                mission.update()
           
            # 將每個類別的總結信息寫入到總結工作表中
            for set_num, count in check_calculation_map.items():
                summary_sheet.append([category, count])
        
        # 保存 Excel 文件
        workbook.save(self.project_path)
        
        print("make_input_excel end")
        return 'make_input_excel end'

    def save_sheet_as_csv(self, sheet_name, csv_file_path):
        # 讀取 Excel 文件中的特定工作表
        df = pd.read_excel(self.project_path, sheet_name=sheet_name)

        # 將 DataFrame 存成 CSV 文件
        df.to_csv(csv_file_path, index=False)
        
        print(f'Saved {sheet_name} as {csv_file_path}')

# handler = MAKE_INFERENCE_DATASET('cats_and_dogs', 'CAT_DOG')
# handler.make_input_excel()
# handler.save_sheet_as_csv('inference_dataset', 'CAT_DOG/temp/inference_dataset.csv')
# handler.save_sheet_as_csv('classification_dataset_Summary', 'CAT_DOG/temp/classification_dataset_detail.csv')